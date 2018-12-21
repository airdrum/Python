import telnetlib
import sys,os
import time
import re 
from statistics import mean , stdev
from _winapi import ReadFile
import xlsxwriter
from openpyxl import load_workbook
import math
import shutil
class FileParser:
        # Initializer / Instance Attributes
    def __init__(self,filepath):
        self.filepath = filepath
    def listFiles(self):
        fileArray = []
        if len(sys.argv) == 2:
            self.filepath  = sys.argv[1]
 
        files = os.listdir(self.filepath)
        for name in files:
            fileArray.append(name)

        return fileArray
    
    def getAbsolutePath(self,file):
        absFile=self.filepath + '\\' +(file)
        return absFile
    
    def readFile(self,fileName):
        file = open(fileName, "r")
        fileStr = file.read()
        numarray=[]
        
        for stra in fileStr.splitlines():
            
            if "quantenna # i=1; while [" in stra:
                print("\nAttenuation ANT0",end=',')
            if "rssi_" in stra:
                print(stra.split("=")[1].replace(' ',''),end=',')
        for stra in fileStr.splitlines():
            
            if "quantenna # i=1; while [" in stra:
                print("\nAttenuation ANT1",end=',')
            if "rssi_1" in stra:
                print(stra.split("=")[1].replace(' ',''),end=',')
        for stra in fileStr.splitlines():
            
            if "quantenna # i=1; while [" in stra:
                print("\nAttenuation ANT2",end=',')
            if "rssi_2" in stra:
                print(stra.split("=")[1].replace(' ',''),end=',')
        for stra in fileStr.splitlines():
            
            if "quantenna # i=1; while [" in stra:
                print("\nAttenuation ANT3",end=',')
            if "rssi_3" in stra:
                print(stra.split("=")[1].replace(' ',''),end=',')


fileName="C:\\Users\\samet.yildiz\\Dropbox\\AirTies\\Throughput Testing\\Air4920 vs Air4921 CH132 20181217"
parser = FileParser(fileName)
filearray=[]
for str in parser.listFiles():
    if ".txt" in str:
        filearray.append(str)
filepath=[]
for inputstr in filearray:
    filepath.append(parser.getAbsolutePath(inputstr))

for readf in filepath:
    print('\n')
    print(readf)
    parser.readFile(readf)

"""parser.parseSimulation(parser,0)
parser.parseSimulation(parser,1)
parser.parseSimulation(parser,2)
parser.closeExcel()"""
#parser.createCombinedAverage(parser)
"""parser.filepath ="E:\\4921\\4921_A1X"
parser.createCombinedAverage(parser)
parser.filepath ="E:\\4921\\4921_A2X"
parser.createCombinedAverage(parser)
parser.closeExcel()
"""
"""wb = load_workbook(filename = 'output.xlsx')
ws1 = wb["Sheet1"] # insert at the end (default)
ws2 = wb["Sheet2"]
ws3 = wb["Sheet3"]
x1 = ws1.cell(row=3, column=2).value
x2 = ws2.cell(row=3, column=2).value
x3 = ws3.cell(row=3, column=2).value
print(x1+x2+x3)
fileName="E:\\4920"
parser = StargateParser(fileName)
print(fileName + '\\' +parser.listFiles()[0])"""